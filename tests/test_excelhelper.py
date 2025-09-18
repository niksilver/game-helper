import pytest

import openpyxl

from openpyxl import Workbook
from excelhelper import ExcelHelper


class TestExcelHelper:


    def test_down(self):
        wb = Workbook()
        xh = ExcelHelper(wb)

        # Check it works with coordinates

        cell = xh.down('C5')
        assert cell.coordinate == 'C6'

        cell = xh.down('E1', 3)
        assert cell.coordinate == 'E4'

        # Check it works with cells

        cell = xh.down(wb.active['C5'])
        assert cell.coordinate == 'C6'

        cell = xh.down(wb.active['E1'], 3)
        assert cell.coordinate == 'E4'


    def test_left(self):
        wb = Workbook()
        xh = ExcelHelper(wb)

        # Check it works with coordinates

        cell = xh.left('C5')
        assert cell.coordinate == 'D5'

        cell = xh.left('E1', 3)
        assert cell.coordinate == 'H1'

        # Check it works with cells

        cell = xh.left(wb.active['C5'])
        assert cell.coordinate == 'D5'

        cell = xh.left(wb.active['E1'], 3)
        assert cell.coordinate == 'H1'


    def test_find(self):
        wb = Workbook()
        xh = ExcelHelper(wb)

        ws = wb.active
        ws['C4'] = 'Here I am!'
        ws['B3'] = 'And again!'

        assert xh.find('Here I am!', 5, 5).coordinate == 'C4'
        assert xh.find('And again!', 5, 5).coordinate == 'B3'

        with pytest.raises(Exception):
            xh.find('Here I am!', 3, 3)


    def test_find_value_beside(self):
        wb = Workbook()
        xh = ExcelHelper(wb)

        ws = wb.active
        ws['C4'] = 'Here I am!'
        ws['D4'] = 333

        assert xh.find_value_beside('Here I am!', 5, 5) == 333

        with pytest.raises(Exception):
            xh.find_value_beside('Here I am!', 3, 3)


    def test_find_non_blank_below(self):
        wb = Workbook()
        xh = ExcelHelper(wb)
        ws = wb.active

        ws['C4'] = 'One'
        ws['C5'] = 'Two'
        ws['C6'] = 'Three'
        ws['C7'] = 'Four'
        # Miss a row here
        ws['C9'] = 'Six'

        cell = xh.find_non_blank_below('C2')

        assert cell.coordinate == 'C4'


    def test_find_non_blank_below_allows_cell(self):
        wb = Workbook()
        xh = ExcelHelper(wb)
        ws = wb.active

        ws['C4'] = 'One'
        ws['C5'] = 'Two'
        ws['C6'] = 'Three'
        ws['C7'] = 'Four'
        # Miss a row here
        ws['C9'] = 'Six'

        cell = xh.find_non_blank_below(ws['C2'])

        assert cell.coordinate == 'C4'


    def test_find_non_blank_below_limits_search(self):
        wb = Workbook()
        hx = ExcelHelper(wb)
        ws = wb.active

        ws['C104'] = 'One'
        ws['C105'] = 'Two'
        ws['C106'] = 'Three'
        ws['C107'] = 'Four'

        with pytest.raises(Exception):
            xh.find_non_blank_below('C2')


    def test_find_values_below(self):
        wb = Workbook()
        xh = ExcelHelper(wb)
        ws = wb.active

        ws['C4'] = 'One'
        ws['C5'] = 'Two'
        ws['C6'] = 'Three'
        ws['C7'] = 'Four'
        # Miss a row here
        ws['C9'] = 'Six'

        arr = xh.find_values_below('C2')

        assert arr == ['One', 'Two', 'Three', 'Four']


    def test_find_values_below_allows_cell(self):
        wb = Workbook()
        xh = ExcelHelper(wb)
        ws = wb.active

        ws['C4'] = 'One'
        ws['C5'] = 'Two'
        ws['C6'] = 'Three'
        ws['C7'] = 'Four'
        # Miss a row here
        ws['C9'] = 'Six'

        arr = xh.find_values_below(ws['C2'])

        assert arr == ['One', 'Two', 'Three', 'Four']


    def test_find_values_below_limits_its_search(self):
        wb = Workbook()
        xh = ExcelHelper(wb)
        ws = wb.active

        ws['C104'] = 'One'
        ws['C105'] = 'Two'
        ws['C106'] = 'Three'
        ws['C107'] = 'Four'

        with pytest.raises(Exception):
            xh.find_values_below('C2')


    def test_put_values_below(self):
        wb = Workbook()
        xh = ExcelHelper(wb)
        xh.put_values_below('D3', ['Aaa', 'Bbb', 'Ccc'])

        ws = wb.active
        assert ws['D4'].value == 'Aaa'
        assert ws['D5'].value == 'Bbb'
        assert ws['D6'].value == 'Ccc'


    def test_put_values_below_allows_cell(self):
        wb = Workbook()
        xh = ExcelHelper(wb)
        ws = wb.active

        xh.put_values_below(ws['D3'], ['Aaa', 'Bbb', 'Ccc'])

        ws = wb.active
        assert ws['D4'].value == 'Aaa'
        assert ws['D5'].value == 'Bbb'
        assert ws['D6'].value == 'Ccc'


    def test_find_values_beside(self):
        wb = Workbook()
        xh = ExcelHelper(wb)
        ws = wb.active

        ws['B2'] = 'Numbers'  # Title of row

        ws['E2'] = 'One'
        ws['F2'] = 'Two'
        ws['G2'] = 'Three'
        ws['H2'] = 'Four'
        # Miss a row here
        ws['J2'] = 'Six'

        arr = xh.find_values_beside('B2')

        assert arr == ['One', 'Two', 'Three', 'Four']


    def test_find_values_beside_allows_cell(self):
        wb = Workbook()
        xh = ExcelHelper(wb)
        ws = wb.active

        ws['B2'] = 'Numbers'  # Title of row

        ws['E2'] = 'One'
        ws['F2'] = 'Two'
        ws['G2'] = 'Three'
        ws['H2'] = 'Four'
        # Miss a row here
        ws['J2'] = 'Six'

        arr = xh.find_values_beside(ws['B2'])

        assert arr == ['One', 'Two', 'Three', 'Four']


    def test_find_values_below_limits_its_search(self):
        wb = Workbook()
        xh = ExcelHelper(wb)
        ws = wb.active

        ws['E104'] = 'One'
        ws['F104'] = 'Two'
        ws['G104'] = 'Three'
        ws['H104'] = 'Four'

        with pytest.raises(Exception):
            xh.find_values_beside('E2')


    def test_put_values_beside(self):
        wb = Workbook()
        xh = ExcelHelper(wb)
        xh.put_values_beside('D3', ['Aaa', 'Bbb', 'Ccc'])

        ws = wb.active
        assert ws['E3'].value == 'Aaa'
        assert ws['F3'].value == 'Bbb'
        assert ws['G3'].value == 'Ccc'


    def test_put_values_beside_allows_cell(self):
        wb = Workbook()
        xh = ExcelHelper(wb)
        ws = wb.active
        xh.put_values_beside(ws['D3'], ['Aaa', 'Bbb', 'Ccc'])

        ws = wb.active
        assert ws['E3'].value == 'Aaa'
        assert ws['F3'].value == 'Bbb'
        assert ws['G3'].value == 'Ccc'


    def test_find_in_table(self):
        wb = Workbook()
        xh = ExcelHelper(wb)
        ws = wb.active

        data = [['Name' , 'Age' , 'Score'],
                ['Alice', 11    , 100],
                ['Bob'  , 12    , 101],
                ['Chris', 13    , 102]]
        for r in range(len(data)):
            for c in range(len(data[r])):
                ws.cell(row = r+8,
                        column = c+4,
                        value = data[r][c])

        # Check it's set up okay

        assert ws['D8'].value == 'Name'

        # Check the method finds data that's there

        assert xh.find_value_in_table('D8', 'Alice', 'Age') == 11
        assert xh.find_value_in_table('D8', 'Bob',   'Age') == 12
        assert xh.find_value_in_table('D8', 'Chris', 'Age') == 13
        assert xh.find_value_in_table('D8', 'Alice', 'Score') == 100
        assert xh.find_value_in_table('D8', 'Bob',   'Score') == 101
        assert xh.find_value_in_table('D8', 'Chris', 'Score') == 102

        # Check the method raises an exception otherwise

        with pytest.raises(Exception) as excinfo:
            xh.find_value_in_table('D8', 'Alice', 'Nonsense')
        assert 'Cannot find' in str(excinfo.value)

        with pytest.raises(Exception) as excinfo:
            xh.find_value_in_table('D8', 'Noone', 'Age')
        assert 'Cannot find' in str(excinfo.value)

        with pytest.raises(Exception):
            xh.find_value_in_table('D8', 'Noone', 'Nonesense')
        assert 'Cannot find' in str(excinfo.value)


    def test_find_in_table_allows_cell(self):
        wb = Workbook()
        xh = ExcelHelper(wb)
        ws = wb.active

        data = [['Name' , 'Age' , 'Score'],
                ['Alice', 11    , 100],
                ['Bob'  , 12    , 101],
                ['Chris', 13    , 102]]
        for r in range(len(data)):
            for c in range(len(data[r])):
                ws.cell(row = r+8,
                        column = c+4,
                        value = data[r][c])

        # Check it's set up okay

        assert ws['D8'].value == 'Name'

        # Check the method finds data that's there

        assert xh.find_value_in_table(ws['D8'], 'Alice', 'Age') == 11


    def test_value_from(self):
        wb = Workbook()
        xh = ExcelHelper(wb)
        ws = wb.active

        data = [['Name' , 'Age' , 'Score'],
                ['Alice', 11    , 100],
                ['Bob'  , 12    , 101],
                ['Chris', 13    , 102]]
        for r in range(len(data)):
            for c in range(len(data[r])):
                ws.cell(row = r+8,
                        column = c+4,
                        value = data[r][c])

        assert xh.value_from('D8', 0, 0) == 'Name'
        assert xh.value_from('D8', 1, 0) == 'Alice'
        assert xh.value_from('D8', 2, 0) == 'Bob'
        assert xh.value_from('D8', 2, 1) == 12
        assert xh.value_from('D8', 2, 2) == 101
