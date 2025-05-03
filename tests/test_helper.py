import pytest

import openpyxl

from openpyxl import Workbook
from excelhelper import ExcelHelper


class TestExcelHelper:


    def test_down(self):
        wb = Workbook()
        xh = ExcelHelper(wb)

        # Check it works with coordinates

        cell = xh.down('C5')
        assert cell.coordinate == 'C6'

        cell = xh.down('E1', 3)
        assert cell.coordinate == 'E4'

        # Check it works with cells

        cell = xh.down(wb.active['C5'])
        assert cell.coordinate == 'C6'

        cell = xh.down(wb.active['E1'], 3)
        assert cell.coordinate == 'E4'


    def test_left(self):
        wb = Workbook()
        xh = ExcelHelper(wb)

        # Check it works with coordinates

        cell = xh.left('C5')
        assert cell.coordinate == 'D5'

        cell = xh.left('E1', 3)
        assert cell.coordinate == 'H1'

        # Check it works with cells

        cell = xh.left(wb.active['C5'])
        assert cell.coordinate == 'D5'

        cell = xh.left(wb.active['E1'], 3)
        assert cell.coordinate == 'H1'


    def test_find(self):
        wb = Workbook()
        xh = ExcelHelper(wb)

        ws = wb.active
        ws['C4'] = 'Here I am!'
        ws['B3'] = 'And again!'

        assert xh.find('Here I am!', 5, 5).coordinate == 'C4'
        assert xh.find('And again!', 5, 5).coordinate == 'B3'

        with pytest.raises(Exception):
            xh.find('Here I am!', 3, 3)


    def test_find_value_beside(self):
        wb = Workbook()
        xh = ExcelHelper(wb)

        ws = wb.active
        ws['C4'] = 'Here I am!'
        ws['D4'] = 333

        assert xh.find_value_beside('Here I am!', 5, 5) == 333

        with pytest.raises(Exception):
            xh.find_value_beside('Here I am!', 3, 3)


    def test_find_non_blank_below(self):
        wb = Workbook()
        xh = ExcelHelper(wb)
        ws = wb.active

        ws['C4'] = 'One'
        ws['C5'] = 'Two'
        ws['C6'] = 'Three'
        ws['C7'] = 'Four'
        # Miss a row here
        ws['C9'] = 'Six'

        cell = xh.find_non_blank_below('C2')

        assert cell.coordinate == 'C4'


    def test_find_non_blank_below_allows_cell(self):
        wb = Workbook()
        xh = ExcelHelper(wb)
        ws = wb.active

        ws['C4'] = 'One'
        ws['C5'] = 'Two'
        ws['C6'] = 'Three'
        ws['C7'] = 'Four'
        # Miss a row here
        ws['C9'] = 'Six'

        cell = xh.find_non_blank_below(ws['C2'])

        assert cell.coordinate == 'C4'


    def test_find_non_blank_below_limits_search(self):
        wb = Workbook()
        hx = ExcelHelper(wb)
        ws = wb.active

        ws['C104'] = 'One'
        ws['C105'] = 'Two'
        ws['C106'] = 'Three'
        ws['C107'] = 'Four'

        with pytest.raises(Exception):
            xh.find_non_blank_below('C2')


    def test_find_values_below(self):
        wb = Workbook()
        xh = ExcelHelper(wb)
        ws = wb.active

        ws['C4'] = 'One'
        ws['C5'] = 'Two'
        ws['C6'] = 'Three'
        ws['C7'] = 'Four'
        # Miss a row here
        ws['C9'] = 'Six'

        arr = xh.find_values_below('C2')

        assert arr == ['One', 'Two', 'Three', 'Four']


    def test_find_values_below_allows_cell(self):
        wb = Workbook()
        xh = ExcelHelper(wb)
        ws = wb.active

        ws['C4'] = 'One'
        ws['C5'] = 'Two'
        ws['C6'] = 'Three'
        ws['C7'] = 'Four'
        # Miss a row here
        ws['C9'] = 'Six'

        arr = xh.find_values_below(ws['C2'])

        assert arr == ['One', 'Two', 'Three', 'Four']


    def test_find_values_below_limits_its_search(self):
        wb = Workbook()
        xh = ExcelHelper(wb)
        ws = wb.active

        ws['C104'] = 'One'
        ws['C105'] = 'Two'
        ws['C106'] = 'Three'
        ws['C107'] = 'Four'

        with pytest.raises(Exception):
            xh.find_values_below('C2')


    def test_put_values_below(self):
        wb = Workbook()
        xh = ExcelHelper(wb)
        xh.put_values_below('D3', ['Aaa', 'Bbb', 'Ccc'])

        ws = wb.active
        assert ws['D4'].value == 'Aaa'
        assert ws['D5'].value == 'Bbb'
        assert ws['D6'].value == 'Ccc'
