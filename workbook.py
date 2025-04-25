import openpyxl

from openpyxl import Workbook as OPWorkbook


# This class is a proxy to an OpenPyxl workbook, using code derived from
# https://stackoverflow.com/questions/29268530/

class Workbook(object):


    @classmethod
    def load(cls, filename):
        wb = openpyxl.reader.excel.load_workbook(filename)
        return Workbook(wb)

    def __init__(self, openpyxl_workbook = None):
        if openpyxl_workbook is None:
            openpyxl_workbook = OPWorkbook()
        self._wb = openpyxl_workbook

    def __getattr__(self, attr):
        if attr in dir(self):
            return getattr(self, attr)
        return getattr(self._wb, attr)

    def __setattr_(self, attr, val):
        if attr in dir(self):
            object.__setattr__(self, attr, val)
        else:
            __setattr__(self._wb, attr, val)

    def __getitem__(self, key):
        return self._wb.__getitem__(key)


    def find(self, value, column = 100, row = 100):
        """
        Find the cell with the given value in the active workbook,
        searching from the top left as far as the given row and column.
        """
        for c in range(1, column+1):
            ws = self.active
            for r in range(1, row+1):
                cell = ws.cell(column = c, row = r)
                if cell.value == value:
                    return cell

        raise LookupError(f'Could not find {value} within {col}umn columns and {row} rows')


    def find_value_beside(self, value, column = 100, row = 100):
        """
        Find the cell with the given value in the active workbook, and return
        the value of the cell to its right.
        Will search from the top left as far as the given row and column.
        """
        for c in range(1, column+1):
            ws = self.active
            for r in range(1, row+1):
                cell = ws.cell(column = c, row = r)
                if cell.value == value:
                    return ws.cell(row = r, column = c+1).value

        raise LookupError(f'Could not find {value} within {column} columns and {row} rows')


    def find_non_blank_below(self, coordinate_or_cell):
        """
        Find the first non-blank cell below the given coordinate or cell.
        """

        coord = coordinate_or_cell
        if type(coordinate_or_cell) is openpyxl.cell.cell.Cell:
            coord = coordinate_or_cell.coordinate

        ws = self.active
        row = ws[coord].row
        col = ws[coord].column

        limit = 100
        max_row = row + limit
        out = []

        # Move down to find where the first value is

        found_value = False
        cell = None

        while not(found_value) and row < max_row:
            row += 1
            cell = ws.cell(row = row, column = col)
            found_value = not(cell.value is None)

        if row == max_row:
            raise(LookupError(f'No values found within {limit} rows of {coord}'))

        return cell

    def find_values_below(self, coordinate_or_cell):
        """
        Find an array of values that sit strictly below the given coordinate or cell.
        Will start at the first non-blank, and stop before the first blank after
        some values.
        """

        ws = self._wb.active
        cell = self.find_non_blank_below(coordinate_or_cell)
        row = cell.row
        col = cell.column
        out = []

        # Move down, adding to our output, until we get an empty cell

        found_blank = False

        while not(found_blank):
            out.append(cell.value)
            row += 1
            cell = ws.cell(row = row, column = col)
            found_blank = (cell.value is None)

        return out


    def put_values_below(self, coordinate_or_cell, array):
        coord = coordinate_or_cell
        if type(coordinate_or_cell) is openpyxl.cell.cell.Cell:
            coord = coordinate_or_cell.coordinate

        ws = self.active
        row = ws[coord].row
        col = ws[coord].column

        for val in array:
            row += 1
            ws.cell(column = col, row = row, value = val)


    def find_values_beside(self, coordinate_or_cell):
        """
        Find an array of values that sit strictly to the right of the given coordinate
        or cell.
        Will start at the first non-blank, and stop before the first blank after
        some values.
        """

        coord = coordinate_or_cell
        if type(coordinate_or_cell) is openpyxl.cell.cell.Cell:
            coord = coordinate_or_cell.coordinate

        ws = self.active
        row = ws[coord].row
        col = ws[coord].column

        limit = 100
        max_col = col + limit
        out = []

        # Move down to find where the first value is

        found_value = False
        cell = None

        while not(found_value) and col < max_col:
            col += 1
            cell = ws.cell(row = row, column = col)
            found_value = not(cell.value is None)

        if col == max_col:
            raise(LookupError(f'No values found within {limit} columns of {coord}'))

        # Move down, adding to our output, until we get an empty cell

        found_blank = False

        while not(found_blank):
            out.append(cell.value)
            col += 1
            cell = ws.cell(row = row, column = col)
            found_blank = (cell.value is None)

        return out


    def put_values_beside(self, coordinate_or_cell, array):
        coord = coordinate_or_cell
        if type(coordinate_or_cell) is openpyxl.cell.cell.Cell:
            coord = coordinate_or_cell.coordinate

        ws = self.active
        row = ws[coord].row
        col = ws[coord].column

        for val in array:
            col += 1
            ws.cell(column = col, row = row, value = val)


    def find_value_in_table(self, coordinate_or_cell, row_name, column_name):
        """
        Given a coordinate or cell, search down for the cell with the row name, across
        for the cell with the column name, and return the value in that cell.
        """

        coord = coordinate_or_cell
        if type(coordinate_or_cell) is openpyxl.cell.cell.Cell:
            coord = coordinate_or_cell.coordinate

        ws = self.active
        cell = ws[coord]
        row = None
        column = None

        for r in range(cell.row, cell.row + 100):
            here = ws.cell(row = r, column = cell.column)
            if here.value == row_name:
                row = r
                break

        for c in range(cell.column, cell.column + 100):
            here = ws.cell(row = cell.row, column = c)
            if here.value == column_name:
                column = c
                break

        if row == None or column == None:
            raise(LookupError(f'Cannot find cell for ({row_name},{column_name}( in table at {coord}'))

        return ws.cell(row = r, column = c).value
