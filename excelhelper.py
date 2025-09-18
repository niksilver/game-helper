import openpyxl

from openpyxl import Workbook


class ExcelHelper(object):
    """
    import openpyxl
    from excelhelper import ExcelHelper

    wb = openpyxl.reader.excel.load_workbook('workbook.xlsx')
    xh = ExcelHelper(wb)
    wb.active = wb['Book 1']

    # Then use the methods of the ExcelHelper.
    """


    def __init__(self, wb_or_filename):
        """
        Call this with an OpenPyXL Workbook object or a string filename.
        """
        if type(wb_or_filename) is str:
            filename = wb_or_filename
            wb       = openpyxl.reader.excel.load_workbook(filename)
            self._wb = wb
        else:
            wb       = wb_or_filename
            self._wb = wb


    @property
    def wb(self):
        """
        The underlying workbook (read-only).
        """
        return self._wb


    def cc(self, coordinate_or_cell):
        """
        Given a coordinate (string) or cell (object) return both the
        coordinate and the cell.
        """
        if type(coordinate_or_cell) is openpyxl.cell.cell.Cell:
            return (coordinate_or_cell.coordinate, coordinate_or_cell)
        else:
            return (coordinate_or_cell, self._wb.active[coordinate_or_cell])


    def down(self, coordinate_or_cell, count = 1):
        """
        Get the cell that is down one (or some other count) from the give cell
        in the active worksheet.
        """

        coord, cell = self.cc(coordinate_or_cell)
        return self._wb.active.cell(row = cell.row + count, column = cell.column)


    def right(self, coordinate_or_cell, count = 1):
        """
        Get the cell that is right one (or some other count) from the give cell
        in the active worksheet.
        """

        coord, cell = self.cc(coordinate_or_cell)
        return self._wb.active.cell(row = cell.row, column = cell.column + count)


    def find(self, value, column = 100, row = 100):
        """
        Find the cell with the given value in the active workbook,
        searching from the top left as far as the given row and column.
        """
        for c in range(1, column+1):
            ws = self._wb.active
            for r in range(1, row+1):
                cell = ws.cell(column = c, row = r)
                if cell.value == value:
                    return cell

        raise LookupError(f'Could not find {value} within {col}umn columns and {row} rows')


    def find_value_beside(self, value, column = 100, row = 100):
        """
        Find the cell with the given value in the active workbook, and return
        the value of the cell to its right.
        Will search from the top left as far as the given row and column.
        """
        for c in range(1, column+1):
            ws = self._wb.active
            for r in range(1, row+1):
                cell = ws.cell(column = c, row = r)
                if cell.value == value:
                    return ws.cell(row = r, column = c+1).value

        raise LookupError(f'Could not find {value} within {column} columns and {row} rows')


    def find_non_blank_below(self, coordinate_or_cell):
        """
        Find the first non-blank cell below the given coordinate or cell.
        """

        coord, cell = self.cc(coordinate_or_cell)

        ws = self._wb.active
        row = cell.row
        col = cell.column

        limit = 100
        max_row = row + limit
        out = []

        # Move down to find where the first value is

        found_value = False
        cell = None

        while not(found_value) and row < max_row:
            row += 1
            cell = ws.cell(row = row, column = col)
            found_value = not(cell.value is None)

        if row == max_row:
            raise(LookupError(f'No values found within {limit} rows of {coord}'))

        return cell


    def find_values_below(self, coordinate_or_cell):
        """
        Find an array of values that sit strictly below the given coordinate or cell.
        Will start at the first non-blank, and stop before the first blank after
        some values.
        """

        ws = self._wb.active
        cell = self.find_non_blank_below(coordinate_or_cell)
        row = cell.row
        col = cell.column
        out = []

        # Move down, adding to our output, until we get an empty cell

        found_blank = False

        while not(found_blank):
            out.append(cell.value)
            row += 1
            cell = ws.cell(row = row, column = col)
            found_blank = (cell.value is None)

        return out


    def put_values_below(self, coordinate_or_cell, array):
        """
        Put an array of values directly below the given coordinate or cell in the
        active workbook.
        """

        coord, cell = self.cc(coordinate_or_cell)

        ws = self._wb.active
        row = cell.row
        col = cell.column

        for val in array:
            row += 1
            ws.cell(column = col, row = row, value = val)


    def find_values_beside(self, coordinate_or_cell):
        """
        Find an array of values that sit strictly to the right of the given coordinate
        or cell.
        Will start at the first non-blank, and stop before the first blank after
        some values.
        """

        coord, cell = self.cc(coordinate_or_cell)

        ws = self._wb.active
        row = cell.row
        col = cell.column

        limit = 100
        max_col = col + limit
        out = []

        # Move down to find where the first value is

        found_value = False
        cell = None

        while not(found_value) and col < max_col:
            col += 1
            cell = ws.cell(row = row, column = col)
            found_value = not(cell.value is None)

        if col == max_col:
            raise(LookupError(f'No values found within {limit} columns of {coord}'))

        # Move down, adding to our output, until we get an empty cell

        found_blank = False

        while not(found_blank):
            out.append(cell.value)
            col += 1
            cell = ws.cell(row = row, column = col)
            found_blank = (cell.value is None)

        return out


    def put_values_beside(self, coordinate_or_cell, array):
        """
        Put an array of values directly to the right of the given coordinate or cell
        in the active workbook.
        """

        coord, cell = self.cc(coordinate_or_cell)

        ws = self._wb.active
        row = cell.row
        col = cell.column

        for val in array:
            col += 1
            ws.cell(column = col, row = row, value = val)


    def find_value_in_table(self, coordinate_or_cell, row_name, column_name):
        """
        Given a coordinate or cell, search down for the cell with the row name, across
        for the cell with the column name, and return the value in that cell.
        """

        coord, cell = self.cc(coordinate_or_cell)
        ws = self._wb.active
        row = None
        column = None

        for r in range(cell.row, cell.row + 100):
            here = ws.cell(row = r, column = cell.column)
            if here.value == row_name:
                row = r
                break

        for c in range(cell.column, cell.column + 100):
            here = ws.cell(row = cell.row, column = c)
            if here.value == column_name:
                column = c
                break

        if row == None or column == None:
            raise(LookupError(f'Cannot find cell for ({row_name},{column_name}( in table at {coord}'))

        return ws.cell(row = r, column = c).value


    def value_from(self, coordinate_or_cell, row = 0, column = 0):
        """
        Given starting coordinate or cell in the active workbook, get the value of the
        cell some number of rows and columns relative to this.
        """

        coord, cell = self.cc(coordinate_or_cell)
        ws = self._wb.active

        return ws.cell(row = cell.row + row, column = cell.column + column).value


    def vertical_table(self, coordinate_or_cell):
        """
        Given a starting cell, which is the first cell of a table header,
        return a table of cells below that. Each element of the table is a
        row of the table. These are the length of the header. The table
        stops just before the first row of all empty cells.
        """

        coord, cell = self.cc(coordinate_or_cell)

        # How many columns in the table?

        header_cell = cell
        cols = 0
        while not(header_cell.value is None):
            cols += 1
            header_cell = self.right(header_cell)

        # Read the rows, stop when we have our first blank

        table         = []
        got_blank_row = False
        start_cell    = self.down(cell)

        while not(got_blank_row):
            found_cell_content = False
            row = []
            for i in range(cols):
                val = self.right(start_cell, i).value
                print(f"Found value '{val}'")
                row.append(val)

                if not(val is None):
                    found_cell_content = True

            table.append(row)
            start_cell = self.down(start_cell)
            got_blank_row = not(found_cell_content)

        # We've appended a blank row, so fix that

        table = table[:-1]

        return table
